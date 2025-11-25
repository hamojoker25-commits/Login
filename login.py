import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
import random
import string
from datetime import datetime
import os

# --- إعدادات الصفحة (تم التصحيح) ---
st.set_page_config(page_title="نظام إدارة المستخدمين", layout="centered")

# اسم ملف الإكسل
EXCEL_FILE = "users_database.xlsx"
MAIN_SHEET_NAME = "All_Users_Data"

# --- دوال المساعدة (Logic Functions) ---

def init_excel():
    """إنشاء ملف الإكسل إذا لم يكن موجوداً"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = MAIN_SHEET_NAME
        # رؤوس الأعمدة في الشيت الرئيسي
        ws.append(["User_Code", "First_Name", "Second_Name", "Email", "Password", "DOB", "Age", "Created_At"])
        wb.save(EXCEL_FILE)

def calculate_age(birth_date):
    """حساب العمر بناءً على تاريخ الميلاد"""
    today = datetime.now().date()
    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
    return age

def generate_user_code():
    """توليد كود: حرف كبير + 5 أرقام عشوائية غير مكررة"""
    letter = random.choice(string.ascii_uppercase)
    digits = random.sample(string.digits, 5)
    code = letter + "".join(digits)
    return code

def save_new_user(first_name, second_name, email, password, dob, age):
    """حفظ المستخدم في الشيت الرئيسي وإنشاء شيت خاص به"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws_main = wb[MAIN_SHEET_NAME]
        
        # التأكد من عدم تكرار الكود
        existing_codes = [row[0] for row in ws_main.iter_rows(min_row=2, values_only=True)] if ws_main.max_row > 1 else []
        
        while True:
            user_code = generate_user_code()
            if user_code not in existing_codes:
                break
        
        # 1. الحفظ في الشيت الرئيسي
        ws_main.append([user_code, first_name, second_name, email, password, dob, age, datetime.now()])
        
        # 2. إنشاء شيت خاص بالمستخدم
        ws_user = wb.create_sheet(title=user_code)
        ws_user.append(["بيانات خاصة بالمستخدم", "ملاحظات", "التاريخ"])
        
        wb.save(EXCEL_FILE)
        return user_code
    except Exception as e:
        st.error(f"حدث خطأ أثناء حفظ البيانات: {e}")
        return None

def verify_login(user_code, password):
    """التحقق من صحة تسجيل الدخول"""
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=MAIN_SHEET_NAME, engine='openpyxl')
        df['User_Code'] = df['User_Code'].astype(str)
        # تحويل الباسورد لـ str للمقارنة الآمنة
        user_row = df[(df['User_Code'] == user_code) & (df['Password'].astype(str) == str(password))]
        
        if not user_row.empty:
            return user_row.iloc[0]
        else:
            return None
    except Exception as e:
        # في حالة عدم وجود الملف بعد، نعتبر قاعدة البيانات فارغة
        return None

# --- واجهة التطبيق (UI) ---

def main():
    init_excel()
    
    st.title("نظام التسجيل المتطور 🚀")

    menu = ["تسجيل الدخول", "إنشاء حساب جديد"]
    
    if 'logged_in' not in st.session_state:
        st.session_state['logged_in'] = False
        st.session_state['user_data'] = None

    if st.session_state['logged_in']:
        # --- صفحة المستخدم ---
        user = st.session_state['user_data']
        
        st.success(f"تم تسجيل الدخول بنجاح! مرحباً بك يا {user['First_Name']}")
        st.divider()
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.info(f"👤 الاسم: {user['First_Name']} {user['Second_Name']}")
        with col2:
            st.warning(f"🔑 الكود الخاص بك: {user['User_Code']}")
        with col3:
            st.info(f"🎂 العمر: {user['Age']} سنة")
            
        st.divider()
        st.subheader("مساحة البيانات الشخصية")
        st.write("هنا ستظهر البيانات الخاصة بك من الشيت الخاص بك.")
        
        try:
            user_sheet_df = pd.read_excel(EXCEL_FILE, sheet_name=str(user['User_Code']), engine='openpyxl')
            st.dataframe(user_sheet_df, use_container_width=True)
        except:
            st.warning("لم يتم العثور على الشيت الخاص بك.")

        if st.button("تسجيل الخروج"):
            st.session_state['logged_in'] = False
            st.session_state['user_data'] = None
            st.rerun()

    else:
        # --- القائمة الجانبية ---
        choice = st.sidebar.selectbox("القائمة", menu)

        if choice == "إنشاء حساب جديد":
            st.header("📝 إنشاء حساب جديد")
            
            with st.form("signup_form"):
                col1, col2 = st.columns(2)
                with col1:
                    f_name = st.text_input("الاسم الأول")
                with col2:
                    s_name = st.text_input("الاسم الثاني")
                
                dob = st.date_input("تاريخ الميلاد", min_value=datetime(1950, 1, 1), max_value=datetime.now())
                email = st.text_input("البريد الإلكتروني")
                pass1 = st.text_input("كلمة المرور", type="password")
                pass2 = st.text_input("تأكيد كلمة المرور", type="password")
                
                submit_signup = st.form_submit_button("تسجيل حساب")
            
            if submit_signup:
                if pass1 != pass2:
                    st.error("❌ كلمات المرور غير متطابقة!")
                elif not f_name or not s_name or not email or not pass1:
                    st.warning("⚠️ يرجى ملء جميع الحقول.")
                else:
                    age = calculate_age(dob)
                    new_code = save_new_user(f_name, s_name, email, pass1, dob, age)
                    
                    if new_code:
                        st.balloons()
                        st.success("✅ تم إنشاء الحساب بنجاح!")
                        st.markdown(f"""
                        ### بياناتك للدخول:
                        - **الاسم:** {f_name} {s_name}
                        - **كود الدخول:** `{new_code}`
                        - **العمر:** {age}
                        """)
        
        elif choice == "تسجيل الدخول":
            st.header("🔐 تسجيل الدخول")
            
            with st.form("login_form"):
                login_code = st.text_input("كود الدخول")
                login_pass = st.text_input("كلمة المرور", type="password")
                submit_login = st.form_submit_button("دخول")
            
            if submit_login:
                user_info = verify_login(login_code, login_pass)
                if user_info is not None:
                    st.session_state['logged_in'] = True
                    st.session_state['user_data'] = user_info
                    st.rerun()
                else:
                    st.error("❌ كود الدخول أو كلمة المرور غير صحيحة.")

if __name__ == '__main__':
    main()
